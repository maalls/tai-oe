#!/usr/bin/env python3
"""
Transform and convert YUASA tariff Excel file to CSV.
Crops rows above the first 'Référence' header in the first column.
"""

from pathlib import Path
import sys
import openpyxl
import csv
import re

from config import get_vendor_config
def main():
    """Transform and convert YUASA spreadsheet to CSV."""
    
    print("🔍 Reading configuration...")
    config = get_vendor_config("yuasa")
    input_file_name = config.get("file_path")
    output_dir_name = config.get("csv_output_dir")
    print(f"📁 YUASA input file: {input_file_name}")
    print(f"📁 CSV output directory: {output_dir_name}")
    input_file = Path(input_file_name)
    output_dir = Path(output_dir_name)
    
    if not input_file.exists():
        print(f"❌ File not found: {input_file}")
        return 1
    
    print(f"📊 Processing {input_file.name}...")
    print(f"   Input: {input_file}")
    print(f"   Output: {output_dir}\n")
    
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Load workbook
        wb = openpyxl.load_workbook(str(input_file))
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            print(f"   📄 Processing sheet: {sheet_name}")
            
            # Find the first row with 'Référence' in the first column
            reference_row = None
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                if row[0].value and str(row[0].value).strip().lower() == 'référence':
                    reference_row = row_idx
                    break
            
            if reference_row is None:
                print(f"      ⚠️  No 'Référence' header found in sheet '{sheet_name}', skipping")
                continue
            
            print(f"      ✓ Found 'Référence' at row {reference_row}")
            print(f"      ✓ Cropping rows 1-{reference_row - 1}")
            
            # Extract data from reference_row onwards
            output_file = output_dir / f"{sheet_name}.csv"
            gammes = []
            with output_file.open('w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # get the
                gamme = None
                
                # Write rows from reference_row onwards
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    
                    if row_idx == reference_row:
                        print(f"      ✓ Writing header row at {row_idx}")
                        # get the row of the next iteration
                        # we overwite none empty values in next_row into row
                        header = ['marque', 'gamme', 'reference', 'tension', 'capacite_10h', 'capacite_20h', 'longueur', 'profondeur', 'hauteur', 'poids', 'carton', 'palette', 'ht']
                        
                        writer.writerow(header)
                    if row_idx > reference_row + 1:

                        # we should stop when we reach 'Conditions commerciales :' in the first column
                        if row[0] and str(row[0]).strip().lower() == 'conditions commerciales :':
                            print(f"      ✓ Stopping at 'Conditions commerciales :' at row {row_idx}")
                            break

                        # we should trim empty rows at the end
                        if all(cell is None for cell in row):
                            continue
                        # we should only write the 11 first columns
                        row = row[:11]

                        # we check if the first cell match the pattern $Gamme ([a-zA-Z0-9_-]+) and we extract the group
                        
                        pattern = r'^Gamme\s+([a-zA-Z0-9_-]+)'
                        match = re.match(pattern, str(row[0]).strip())
                        if (match):
                            gamme = match.group(1)
                            gammes.append({ 
                                'reference': gamme,
                                'description': row[0]
                            })
                            continue
                        
                        if not gamme:
                            print(f"      ⚠️  No 'Gamme' found before data row at {row_idx}, skipping row")
                            raise Exception("No 'Gamme' found before data row")
                        
                        # add gamme to the row
                        row = list(row)
                        row.insert(0, gamme)

                        # Add the vendor code to the row
                        # write the row
                        row.insert(0, 'YUASA')
                       
                        writer.writerow(row)
            
            print(f"      ✅ Saved: {output_file.name}")
        
        print("\n✨ Transform complete!")
        # storing gammes in a second csv called gammes.csv
        gammes_file = output_dir / "gammes.csv"
        with gammes_file.open('w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['reference', 'description'])
            for gamme in gammes:
                writer.writerow([gamme['reference'], gamme['description']]) 

        print(f"   ✅ Gammes saved: {gammes_file.name}")
        return 0
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
